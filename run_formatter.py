import pandas as pd
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_calendar_sheet(wb, month_date, schedule_df, rotation_periods=None):
    print(f"create_calendar_sheet called for {month_date.strftime('%B %Y')}")
    print(f"schedule_df shape: {schedule_df.shape}")
    
    # Create new sheet with month name
    month_name = month_date.strftime("%B %Y")
    ws = wb.create_sheet(title=month_name)
    
    # Write month and year in row 1
    ws.cell(row=1, column=1, value=month_name)
    
    # Column mappings for days of week (each day gets 2 columns)
    day_columns = {
        'Sunday': ['A', 'B'],
        'Monday': ['C', 'D'],
        'Tuesday': ['E', 'F'],
        'Wednesday': ['G', 'H'],
        'Thursday': ['I', 'J'],
        'Friday': ['K', 'L'],
        'Saturday': ['M', 'N']
    }
    
    # Set column widths
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 5

    # Define border styles
    thin_border = Side(style='thin')
    day_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )

    # Write day headers in row 2
    for day, cols in day_columns.items():
        col_idx = ord(cols[0]) - ord('A') + 1
        for offset in range(2):
            cell = ws.cell(row=2, column=col_idx + offset)
            cell.value = day if offset == 0 else ""
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Define border styles
    thin_border = Side(style='thin')
    
    # Colors
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')     # Light gray
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')    # Light green
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Light yellow
    
    # Get first and last day of the month
    first_of_month = datetime(month_date.year, month_date.month, 1)
    _, num_days = calendar.monthrange(month_date.year, month_date.month)
    last_of_month = datetime(month_date.year, month_date.month, num_days)

    # Find the first Sunday before or on the 1st
    first_sunday = first_of_month
    while first_sunday.weekday() != 6:  # 6 = Sunday
        first_sunday -= timedelta(days=1)
    # Find the last Saturday after or on the last day
    last_saturday = last_of_month
    while last_saturday.weekday() != 5:  # 5 = Saturday
        last_saturday += timedelta(days=1)

    # Build list of all days to display
    num_days_display = (last_saturday - first_sunday).days + 1
    all_days = [first_sunday + timedelta(days=i) for i in range(num_days_display)]

    # Build weeks
    calendar_weeks = [all_days[i:i+7] for i in range(0, len(all_days), 7)]

    # Write calendar
    for week_num, week in enumerate(calendar_weeks):
        base_row = 3 + (week_num * 8)  # Each week takes 8 rows
        
        # Add week separator border if not first week
        if week_num > 0:
            for col in range(1, 16):  # A through O
                cell = ws.cell(row=base_row - 1, column=col)
                cell.border = Border(bottom=Side(style='thin'))
        
        # Color the rows for each week
        for col in range(1, 16):  # A through O (including column O)
            # Rows 4-5 (gray)
            ws.cell(row=base_row + 3, column=col).fill = gray_fill
            ws.cell(row=base_row + 4, column=col).fill = gray_fill
            # Row 7 (green)
            ws.cell(row=base_row + 6, column=col).fill = green_fill
            # Row 8 (yellow)
            ws.cell(row=base_row + 7, column=col).fill = yellow_fill
        
        # Add labels in column O for each week
        on_call_label = ws.cell(row=base_row + 3, column=15, value="On Call")
        on_call_label.font = Font(color='00B0F0')  # Light blue color, not bold
        ws.cell(row=base_row + 4, column=15, value="Intern")  # Add intern label
        ws.cell(row=base_row + 6, column=15, value="Supervisor")  # Supervisor label above backup
        ws.cell(row=base_row + 7, column=15, value="Backup")
        
        for weekday, date in enumerate(week):
            cols = list(day_columns.values())[weekday]
            col_idx = ord(cols[0]) - ord('A') + 1
            # Write day number (only in first column)
            day_cell = ws.cell(row=base_row, column=col_idx)
            day_cell.value = date.day
            day_cell.alignment = Alignment(horizontal='center')
            # Italicize if not in current month
            if date.month != month_date.month:
                day_cell.font = Font(italic=True)
            
            # Check if this date is a rotation switch date
            if rotation_periods:
                for rotation in rotation_periods:
                    # Convert both dates to date objects for comparison
                    switch_date = rotation['switch_date']
                    if hasattr(switch_date, 'date'):
                        switch_date = switch_date.date()
                    
                    current_date = date
                    if hasattr(current_date, 'date'):
                        current_date = current_date.date()
                    
                    if switch_date == current_date:
                        # Add SWITCH marker in the cell below the date number
                        switch_cell = ws.cell(row=base_row + 1, column=col_idx)
                        switch_cell.value = "SWITCH"
                        switch_cell.font = Font(bold=True, color='FF0000')  # Bold red text
                        switch_cell.alignment = Alignment(horizontal='center')
                        switch_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background
                        break
            # Get schedule for this day (including overflow days)
            date_str = date.strftime("%Y-%m-%d")
            day_schedule = schedule_df[schedule_df['Date'] == date_str]
            
            if not day_schedule.empty:
                # Write Call person (row 4, first column only)
                call_cell = ws.cell(row=base_row + 3, column=col_idx)
                call_cell.value = day_schedule.iloc[0].get('Call', '')
                call_cell.alignment = Alignment(horizontal='center')
                # Write Intern person (row 5, first column only)
                intern_cell = ws.cell(row=base_row + 4, column=col_idx)
                intern_cell.value = day_schedule.iloc[0].get('Intern', '')
                intern_cell.alignment = Alignment(horizontal='center')
                # Write Supervisor (row 7, first column only)
                if 'Supervisor' in day_schedule.columns:
                    supervisor_cell = ws.cell(row=base_row + 6, column=col_idx)
                    supervisor_cell.value = day_schedule.iloc[0].get('Supervisor', '')
                    supervisor_cell.alignment = Alignment(horizontal='center')
                # Write Backup person (row 8, first column only)
                backup_cell = ws.cell(row=base_row + 7, column=col_idx)
                backup_cell.value = day_schedule.iloc[0].get('Backup', '')
                backup_cell.alignment = Alignment(horizontal='center')
                
                # Style the Call person (On Call) in light blue and bold
                call_cell.font = Font(bold=True, color='00B0F0')  # Light blue and bold
                
                # Add borders around the entire 2x8 day block
                for row_offset in range(8):
                    for col_offset in range(2):
                        current_cell = ws.cell(row=base_row + row_offset, column=col_idx + col_offset)
                        has_left = (col_offset == 0)
                        has_right = (col_offset == 1)
                        has_top = (row_offset == 0)
                        has_bottom = (row_offset == 7)
                        current_cell.border = Border(
                            left=thin_border if has_left else None,
                            right=thin_border if has_right else None,
                            top=thin_border if has_top else None,
                            bottom=thin_border if has_bottom else None
                        )
                
                # Ensure the right border of the entire day block is visible on all rows
                # This fixes the missing vertical borders on the right edge of each day
                for row_offset in range(8):
                    rightmost_cell = ws.cell(row=base_row + row_offset, column=col_idx + 1)
                    current_border = rightmost_cell.border
                    rightmost_cell.border = Border(
                        left=current_border.left,
                        right=Side(style='medium'),  # Make right border more visible
                        top=current_border.top,
                        bottom=current_border.bottom
                    )
                
                # Remove any week separator borders that might interfere with day blocks
                if week_num > 0:
                    for col_offset in range(2):
                        separator_cell = ws.cell(row=base_row - 1, column=col_idx + col_offset)
                        separator_cell.border = Border()
    return ws

def format_schedule(schedule_df, call_distribution_df=None, golden_weekends_data=None, 
                   soft_constraint_results=None, running_totals_df=None, rotation_periods=None,
                   call_by_rotation_df=None):
    """
    Create comprehensive Excel workbook with calendar sheets and data sheets.
    
    Args:
        schedule_df: Main schedule DataFrame with Date, Call, Backup, Intern, Supervisor
        call_distribution_df: Call/backup distribution statistics
        golden_weekends_data: Dict of {resident: count}
        soft_constraint_results: DataFrame with soft constraint analysis
        running_totals_df: DataFrame with cumulative totals across blocks
        rotation_periods: List of dicts with switch_date and rotation_name for marking switch dates
        call_by_rotation_df: DataFrame with call shift counts by rotation for PGY2/PGY3
    """
    # Ensure schedule_df has proper date format
    schedule_df = schedule_df.copy()
    schedule_df['Date'] = pd.to_datetime(schedule_df['Date'])

    # Create workbook
    wb = Workbook()

    # Get all unique months in the schedule
    start = schedule_df['Date'].min()
    end = schedule_df['Date'].max()
    if pd.isna(start) or pd.isna(end):
        raise ValueError("Cannot generate calendar: schedule has no valid dates.")
    
    # Get the first day of the month containing the start date
    start_month_start = datetime(start.year, start.month, 1)
    # Get the first day of the month containing the end date
    end_month_start = datetime(end.year, end.month, 1)
    # Generate all month starts from start month to end month (inclusive)
    months = pd.date_range(start_month_start, end_month_start, freq='MS')
    
    # Create calendar sheets first (visual formatting)
    for month in months:
        # Get first and last day of the month
        first_of_month = datetime(month.year, month.month, 1)
        _, num_days = calendar.monthrange(month.year, month.month)
        last_of_month = datetime(month.year, month.month, num_days)
        # Find the first Sunday before or on the 1st
        first_sunday = first_of_month
        while first_sunday.weekday() != 6:  # 6 = Sunday
            first_sunday -= timedelta(days=1)
        # Find the last Saturday after or on the last day
        last_saturday = last_of_month
        while last_saturday.weekday() != 5:  # 5 = Saturday
            last_saturday += timedelta(days=1)
        # Get all assignments for the full calendar grid (including overflow days)
        mask = (schedule_df['Date'] >= first_sunday) & (schedule_df['Date'] <= last_saturday)
        month_df = schedule_df[mask]
        create_calendar_sheet(wb, month, month_df, rotation_periods)

    # Create data sheets (simple data only)
    if call_distribution_df is not None:
        create_call_distribution_sheet(wb, call_distribution_df)
    
    if golden_weekends_data is not None:
        create_golden_weekends_sheet(wb, golden_weekends_data)
    
    create_raw_schedule_sheet(wb, schedule_df)
    
    if soft_constraint_results is not None:
        create_soft_constraints_sheet(wb, soft_constraint_results)
    
    if running_totals_df is not None:
        create_running_totals_sheet(wb, running_totals_df)
    
    if call_by_rotation_df is not None:
        create_call_by_rotation_sheet(wb, call_by_rotation_df)

    # Remove the default sheet
    wb.remove(wb['Sheet'])
    
    return wb

def create_call_distribution_sheet(wb, call_distribution_df):
    """Create simple call distribution data sheet"""
    ws = wb.create_sheet(title="Call Distribution")
    
    # Write headers
    for col_idx, col_name in enumerate(call_distribution_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, (_, row) in enumerate(call_distribution_df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return ws

def create_golden_weekends_sheet(wb, golden_weekends_data):
    """Create golden weekends data sheet - handles both rotation-based and total formats"""
    ws = wb.create_sheet(title="Golden Weekends")
    
    # Check if data is rotation-based (nested dict) or total-based (flat dict)
    if golden_weekends_data and isinstance(list(golden_weekends_data.values())[0], dict):
        # Rotation-based data format
        ws.cell(row=1, column=1, value="Rotation")
        ws.cell(row=1, column=2, value="Resident")
        ws.cell(row=1, column=3, value="Golden Weekends")
        
        row_idx = 2
        for rotation_name, residents_data in golden_weekends_data.items():
            for resident, count in residents_data.items():
                ws.cell(row=row_idx, column=1, value=rotation_name)
                ws.cell(row=row_idx, column=2, value=resident)
                ws.cell(row=row_idx, column=3, value=count)
                row_idx += 1
    else:
        # Total-based data format (fallback)
        ws.cell(row=1, column=1, value="Resident")
        ws.cell(row=1, column=2, value="Golden Weekends")
        
        for row_idx, (resident, count) in enumerate(golden_weekends_data.items(), 2):
            ws.cell(row=row_idx, column=1, value=resident)
            ws.cell(row=row_idx, column=2, value=count)
    
    return ws

def create_raw_schedule_sheet(wb, schedule_df):
    """Create simple raw schedule data sheet"""
    ws = wb.create_sheet(title="Raw Schedule")
    
    # Write headers
    for col_idx, col_name in enumerate(schedule_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, (_, row) in enumerate(schedule_df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            # Convert datetime to string for Excel compatibility
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return ws

def create_soft_constraints_sheet(wb, soft_constraint_results):
    """Create simple soft constraints results sheet"""
    ws = wb.create_sheet(title="Soft Constraints")
    
    # Write headers
    for col_idx, col_name in enumerate(soft_constraint_results.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, (_, row) in enumerate(soft_constraint_results.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            # Convert dates to strings for Excel compatibility
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return ws

def create_running_totals_sheet(wb, running_totals_df):
    """Create simple running totals data sheet"""
    ws = wb.create_sheet(title="Running Totals")
    
    # Write headers
    for col_idx, col_name in enumerate(running_totals_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, (_, row) in enumerate(running_totals_df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return ws

def create_call_by_rotation_sheet(wb, call_by_rotation_df):
    """Create call shifts by rotation data sheet for PGY2 and PGY3"""
    ws = wb.create_sheet(title="Call by Rotation")
    
    # Write headers
    for col_idx, col_name in enumerate(call_by_rotation_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, (_, row) in enumerate(call_by_rotation_df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return ws

def create_merged_calendar_sheet(wb, prev_month, current_month, schedule_df):
    """Create a calendar sheet that includes the last week of previous month and current month"""
    # Use current month for sheet name
    month_name = current_month.strftime("%B %Y")
    ws = wb.create_sheet(title=month_name)
    
    # Write month and year in row 1
    ws.cell(row=1, column=1, value=month_name)
    
    # Column mappings for days of week (each day gets 2 columns)
    day_columns = {
        'Sunday': ['A', 'B'],
        'Monday': ['C', 'D'],
        'Tuesday': ['E', 'F'],
        'Wednesday': ['G', 'H'],
        'Thursday': ['I', 'J'],
        'Friday': ['K', 'L'],
        'Saturday': ['M', 'N']
    }
    
    # Set column widths
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 5

    # Define border styles
    thin_border = Side(style='thin')
    
    # Write day headers in row 2
    for day, cols in day_columns.items():
        col_idx = ord(cols[0]) - ord('A') + 1
        for offset in range(2):
            cell = ws.cell(row=2, column=col_idx + offset)
            cell.value = day if offset == 0 else ""
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Colors
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Get the earliest date in the schedule
    start_date = schedule_df['Date'].min()
    
    # Calculate the first Sunday of our calendar view
    first_sunday = start_date - pd.Timedelta(days=start_date.weekday() + 1)
    if first_sunday.month != start_date.month:
        first_sunday = start_date - pd.Timedelta(days=start_date.weekday())
    
    # Get number of days to show
    last_date = schedule_df['Date'].max()
    num_days = (last_date - first_sunday).days + 1
    
    # Calculate calendar weeks
    current_week = []
    calendar_weeks = []
    current_date = first_sunday
    
    for _ in range(num_days):
        current_week.append(current_date)
        if len(current_week) == 7:
            calendar_weeks.append(current_week)
            current_week = []
        current_date += pd.Timedelta(days=1)
    
    # Add remaining days if any
    if current_week:
        while len(current_week) < 7:
            current_week.append(current_date)
            current_date += pd.Timedelta(days=1)
        calendar_weeks.append(current_week)
    
    # Write calendar
    for week_num, week in enumerate(calendar_weeks):
        base_row = 3 + (week_num * 8)  # Each week takes 8 rows
        
        # Add week separator border if not first week
        if week_num > 0:
            for col in range(1, 16):  # A through O
                cell = ws.cell(row=base_row - 1, column=col)
                cell.border = Border(bottom=Side(style='thin'))
        
        # Color the rows for each week
        for col in range(1, 16):  # A through O (including column O)
            # Rows 4-5 (gray)
            ws.cell(row=base_row + 3, column=col).fill = gray_fill
            ws.cell(row=base_row + 4, column=col).fill = gray_fill
            
            # Row 7 (green)
            ws.cell(row=base_row + 6, column=col).fill = green_fill
            
            # Row 8 (yellow)
            ws.cell(row=base_row + 7, column=col).fill = yellow_fill
        
        # Add labels in column O for each week
        on_call_label = ws.cell(row=base_row + 3, column=15, value="On Call")
        on_call_label.font = Font(color='00B0F0')  # Light blue color, not bold
        ws.cell(row=base_row + 4, column=15, value="Intern")
        ws.cell(row=base_row + 6, column=15, value="Supervisor")
        ws.cell(row=base_row + 7, column=15, value="Backup")
        
        for weekday, date in enumerate(week):
            print(f"Processing date: {date.strftime('%Y-%m-%d')}")
            # Get the starting column for this day
            cols = list(day_columns.values())[weekday]
            col_idx = ord(cols[0]) - ord('A') + 1
            

            
            # Write day number (only in first column)
            day_cell = ws.cell(row=base_row, column=col_idx)
            day_cell.value = date.day
            day_cell.alignment = Alignment(horizontal='center')
            
            # Style differently if from previous month
            if date.month != current_month.month:
                day_cell.font = Font(italic=True, color='808080')
            
            # Get schedule for this day
            date_str = date.strftime("%Y-%m-%d")
            day_schedule = schedule_df[schedule_df['Date'] == date_str]
            
            # Debug: Print what we found for this date
            if not day_schedule.empty:
                print(f"Date {date_str}: Found {len(day_schedule)} rows")
                print(f"Columns: {list(day_schedule.columns)}")
                print(f"First row data: {day_schedule.iloc[0].to_dict()}")
            else:
                print(f"Date {date_str}: No schedule data found")
            
            if not day_schedule.empty:
                # Write Call person (row 4, first column only)
                call_cell = ws.cell(row=base_row + 3, column=col_idx)
                call_cell.value = day_schedule.iloc[0]['Call']
                call_cell.alignment = Alignment(horizontal='center')
                
                # Write Intern person (row 5, first column only)
                intern_cell = ws.cell(row=base_row + 4, column=col_idx)
                intern_cell.value = day_schedule.iloc[0]['Intern']
                intern_cell.alignment = Alignment(horizontal='center')
                
                # Write Supervisor (row 7, first column only)
                if 'Supervisor' in day_schedule.columns:
                    supervisor_cell = ws.cell(row=base_row + 6, column=col_idx)
                    supervisor_cell.value = day_schedule.iloc[0]['Supervisor']
                    supervisor_cell.alignment = Alignment(horizontal='center')
                
                # Write Backup person (row 8, first column only)
                backup_cell = ws.cell(row=base_row + 7, column=col_idx)
                backup_cell.value = day_schedule.iloc[0]['Backup']
                backup_cell.alignment = Alignment(horizontal='center')
                
                # Style the Call person (On Call) in light blue and bold
                call_cell.font = Font(bold=True, color='00B0F0')  # Light blue and bold
                
                # Add borders around the entire 2x8 day block
                for row_offset in range(8):
                    for col_offset in range(2):
                        current_cell = ws.cell(row=base_row + row_offset, column=col_idx + col_offset)
                        
                        # Determine which borders this cell should have
                        has_left = (col_offset == 0)  # Left border for first column of day
                        has_right = (col_offset == 1)  # Right border for second column of day
                        has_top = (row_offset == 0)   # Top border for first row of day
                        has_bottom = (row_offset == 7) # Bottom border for last row of day
                        
                        # Create border object with all required sides
                        current_cell.border = Border(
                            left=thin_border if has_left else None,
                            right=thin_border if has_right else None,
                            top=thin_border if has_top else None,
                            bottom=thin_border if has_bottom else None
                        )
                        
                        # Ensure the last row has a more visible bottom border
                        if row_offset == 7:  # Last row of the day
                            current_cell.border = Border(
                                left=current_cell.border.left,
                                right=current_cell.border.right,
                                top=current_cell.border.top,
                                bottom=Side(style='medium')  # Make bottom border more visible
                            )
                            
                            # Set row height to ensure bottom border is visible
                            ws.row_dimensions[base_row + row_offset].height = 20
            
            # Ensure the right border of the entire day block is visible on all rows
            # This fixes the missing vertical borders on the right edge of each day
            for row_offset in range(8):
                rightmost_cell = ws.cell(row=base_row + row_offset, column=col_idx + 1)
                current_border = rightmost_cell.border
                rightmost_cell.border = Border(
                    left=current_border.left,
                    right=Side(style='medium'),  # Make right border more visible
                    top=current_border.top,
                    bottom=current_border.bottom
                )
            
            # Remove any week separator borders that might interfere with day blocks
            if week_num > 0:
                for col_offset in range(2):
                    separator_cell = ws.cell(row=base_row - 1, column=col_idx + col_offset)
                    separator_cell.border = Border()
                    

    
    return ws

# Only keep the function definitions, remove the file saving code
if __name__ == '__main__':
    pass 